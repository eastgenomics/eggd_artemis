"""General IO related functions"""

from copy import deepcopy
import json
import re

import dxpy
from mergedeep import merge, Strategy
from openpyxl.styles import DEFAULT_FONT, Font, Protection
import pandas as pd

from .defaults import build_37_urls, build_38_urls
from .util_functions import (
    check_session_track_order,
    filter_reference_tracks,
    set_order_map,
)


def make_cnv_session(
    sample,
    session_file_name,
    bam_url,
    bai_url,
    bed_url,
    seg_url,
    excluded_url,
    targets_url,
    job_output,
    expiry_date,
    build,
    select_tracks,
) -> str:
    """
    Create a session file for IGV

    Args:
        sessions_list (list): list of session files to upload in the end
        sample (string): sample name
        session_file_name (string): str to add to end of session file name
        bam_url (string): URL of the bam file
        bai_url (string): URL of the bai file
        bed_url (string): URL of the bed file
        seg_url (string): URL of the seg file
        excluded_url (string): URL of the excluded regions file
        targets_url (string): URL of the targets file
        job_output (str) : output folder set for job
        expiry_date (string): date of expiry of file created
        build (int): genome build to add reference files to the
            session file for
        select_tracks (str): comma separated string of IGV reference
            tracks to select

    Returns:
        session_file (string): IGV session file URL

    Raises:
        RuntimeError: Raised if invalid build param provided
    """
    order_map = set_order_map()

    with open(
        "/home/dnanexus/cnv-template.json", encoding="utf-8", mode="r"
    ) as fh:
        template = json.load(fh)

    if build == 37:
        reference_urls = build_37_urls
    elif build == 38:
        reference_urls = build_38_urls
    else:
        raise RuntimeError(
            f"Invalid build param given ({build}), must be one of 37 or 38"
        )

    # check if we're filtering down the reference tracks by those provided
    if select_tracks:
        reference_urls["tracks"] = filter_reference_tracks(
            select_tracks=select_tracks,
            reference_tracks=reference_urls["tracks"],
        )

    template = merge(template, reference_urls, strategy=Strategy.ADDITIVE)

    template["tracks"] = check_session_track_order(
        session_tracks=template["tracks"]
    )

    template_copy = deepcopy(template)

    template_copy["tracks"] = []

    # Order 6 - Visualisation bed for CNV calls
    order_map["6"]["url"] = bed_url

    # Order 7 - Patient bam and index
    order_map["7"]["name"] = sample
    order_map["7"]["url"] = bam_url
    order_map["7"]["indexURL"] = bai_url

    # Order 8 - Variant seg file
    order_map["8"]["url"] = seg_url

    # Order 9 - Excluded regions from CNV calling bed
    order_map["9"]["url"] = excluded_url

    # Order 10 - Target regions bed
    order_map["10"]["url"] = targets_url

    # Map urls to template
    for track in template["tracks"]:
        if track["order"] in [6, 7, 8, 9, 10]:
            order = str(track["order"])
            track["url"] = order_map[order]["url"]

        if track["order"] == 7:
            track["indexURL"] = order_map[order]["indexURL"]

        template_copy["tracks"].append(track)

    output_name = f"{sample}_{session_file_name}_igv.json"

    with open(output_name, "w") as outfile:
        json.dump(template_copy, outfile, indent=4)

    session_file = dxpy.upload_local_file(
        output_name,
        folder=f"{job_output}/igv_sessions",
        tags=[expiry_date],
        wait_on_close=True,
    )

    return session_file.get_id()


def read_excluded_regions_to_df(file_id, project) -> pd.DataFrame:
    """
    Read in excluded regions file to a pandas dataframe.

    Args:
        file_id (str): DNA nexus file ID of excluded regions file.
        project (str): DNA nexus project ID of file.

    Raises:
        AssertionError if not all required file headers are present in the
          resulting dataframe.

    Returns:
        pd.DataFrame: file read in as pandas dataframe.
    """
    required_headers = [
        "Chrom",
        "Start",
        "End",
        "Length",
        "Gene_Symbol",
        "HGNC_ID",
        "Transcript",
        "Exon",
    ]

    file = dxpy.open_dxfile(
        file_id,
        project=project,
        mode="r",
    )

    # letters a -> i used as columns names to aid concatenation of this df when
    # forming the output df
    df = pd.read_csv(
        file,
        sep="\t",
        names=["b", "c", "d", "e", "f", "g", "h", "i"],
        header=None,
    )

    headers = df.iloc[0, :].tolist()
    missing = set(required_headers) - set(headers)

    assert not missing, f"{file_id} is missing {missing} in its headers"

    df.insert(0, "a", ["CNV excluded regions"] + [None] * (len(df) - 1))

    return df


def write_output_file(
    sample_outputs,
    today,
    expiry_date,
    multiqc_url,
    qc_url,
    project_name,
    lock_cells,
) -> str:
    """
    Writes output xlsx file with all download URLs

    Parameters
    ----------
    sample_outputs : dict
        generated outputs (i.e. URLs, text and dataframes) for each sample
    today : str
        today date
    expiry_date : str
        date of URL experation
    multiqc_url : str
        download URL for multiQC report
    qc_url : str
        download URL for QC report
    project_name: str
        DNAnexus project name
    lock_cells : boolean
        determines whether to lock any populated cells for editing

    Returns
    -------
    output_xlsx_file (str): filename of the written xlsx file

    Outputs
    -------
    xlsx file
    """
    print("Writing output file")

    sample_count = str(len(sample_outputs.keys()))

    multiqc_url = f'=HYPERLINK("{multiqc_url}", "{multiqc_url}")'
    if qc_url.startswith("http"):
        qc_url = f'=HYPERLINK("{qc_url}", "{qc_url}")'

    # Empty dicts {} are added to create empty rows in order to better
    # organise the output df and resulting spreadsheet
    df = pd.DataFrame(
        [
            {"a": "Run:", "b": project_name},
            {"a": "Number of samples in this file:", "b": sample_count},
            {},
            {"a": "Date Created:", "b": today},
            {"a": "Expiry Date:", "b": expiry_date},
            {},
            {"a": "Run Level Files"},
            {"a": "MultiQC report", "b": multiqc_url},
            {"a": "QC Status Report", "b": qc_url},
            {},
            {},
            {"a": "Per Sample Files"},
            {},
        ]
    )

    sample_order = sorted(sample_outputs.keys())

    for sample in sample_order:
        outputs = sample_outputs.get(sample)

        df = df.append({"a": sample}, ignore_index=True)

        # Fields we need once per sample
        sample_level_urls = {
            "bam_url": "Alignment BAM",
            "bai_url": "Alignment BAI",
        }

        # Fields we need once per report
        output_fields = {
            "coverage_url": "Coverage report",
            "coverage_summary": "Coverage summary",
            "SNV count": "SNV count post-filtering",
            "snv_url": "Small variant report",
            "CNV count": "CNV count",
            "cnv_url": "CNV variant report",
            "cnv_session_url": "CNV IGV Session",
            "cnv_excluded_regions_df": "CNV excluded regions",
        }

        for clinical_indication in outputs["clinical_indications"]:
            file_data = outputs["clinical_indications"][clinical_indication]

            # If we know clinical indication add this as a header
            if clinical_indication != "Unknown":
                df = df.append({"a": clinical_indication}, ignore_index=True)

            # Add fields for any SNV files first for that clinical indication
            for field, label in output_fields.items():
                for snv_data in file_data.get("SNV", {}):
                    if snv_data.get(field):
                        df = df.append(
                            {"a": label, "b": snv_data.get(field)},
                            ignore_index=True,
                        )

                # Then add fields for CNV fields for that clinical indication
                for cnv_data in file_data.get("CNV", {}):
                    # For excluded regions the output is always present but
                    # could be text or a df which affects how it is added to
                    # output df
                    if field == "cnv_excluded_regions_df":
                        if isinstance(cnv_data.get(field), str):
                            df = df.append(
                                {"a": label, "b": cnv_data.get(field)},
                                ignore_index=True,
                            )
                        else:
                            df = pd.concat(
                                [df, cnv_data.get(field)], ignore_index=True
                            )
                    else:
                        if cnv_data.get(field):
                            df = df.append(
                                {"a": label, "b": cnv_data.get(field)},
                                ignore_index=True,
                            )
        # Add BAM and BAI URLs for the sample
        for field, label in sample_level_urls.items():
            if outputs.get(field):
                if field == "bam_url":
                    df = df.append({}, ignore_index=True)

                df = df.append(
                    {"a": label, "b": outputs.get(field)}, ignore_index=True
                )

        df = df.append({}, ignore_index=True)
        df = df.append({}, ignore_index=True)

    writer = pd.ExcelWriter(f"{project_name}_{today}.xlsx", engine="openpyxl")
    df.to_excel(writer, index=False, header=False)

    # set column widths
    sheet = writer.sheets["Sheet1"]
    sheet.column_dimensions["A"].width = 55
    sheet.column_dimensions["B"].width = 11
    sheet.column_dimensions["C"].width = 11
    sheet.column_dimensions["D"].width = 11
    sheet.column_dimensions["E"].width = 11
    sheet.column_dimensions["F"].width = 16
    sheet.column_dimensions["G"].width = 12
    sheet.column_dimensions["H"].width = 16
    sheet.column_dimensions["I"].width = 11

    sheet["A1"].font = Font(bold=True, name=DEFAULT_FONT.name)
    sheet["A2"].font = Font(bold=True, name=DEFAULT_FONT.name)
    sheet["A7"].font = Font(bold=True, name=DEFAULT_FONT.name)
    sheet["A12"].font = Font(bold=True, name=DEFAULT_FONT.name)

    # If lock_cells=True we're protecting populated cells from editing
    # openpyxl is silly and you need to lock a whole sheet then unlock
    # specific cells - so to do the opposite we unlock the number of rows/cols
    # used (via max rows/cols) and add a buffer of 1000/100 rows/cols so that
    # surrounding cells are able to be edited. We then lock any populated cells
    # Also set format so that we can still resize rows/columns if protected
    if lock_cells:
        print("lock_cells=True, locking any populated Excel cells")
        sheet.protection.sheet = True
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False
        sheet.protection.formatCells = False

        unlock_rows = sheet.max_row + 1000
        unlock_cols = sheet.max_column + 100

        for row_no in range(1, unlock_rows):
            for col_no in range(1, unlock_cols):
                sheet.cell(row=row_no, column=col_no).protection = Protection(
                    locked=False
                )
    else:
        print("lock_cells=False, all Excel cells will be editable")

    # Make sample IDs and any clinical indication names bold
    # Lock any cells in column A for editing if they're populated
    for cell in sheet.iter_rows(max_col=1):
        if cell[0].value:
            if re.match(r"X[\d]+", cell[0].value):
                sheet[cell[0].coordinate].font = Font(
                    bold=True, name=DEFAULT_FONT.name
                )
            elif re.match(
                r"[a-zA-Z0-9]+-[a-zA-Z0-9]+-[a-zA-Z0-9]+-[a-zA-Z0-9]+-[MFU]-[a-zA-Z0-9]+",
                cell[0].value,
            ):
                sheet[cell[0].coordinate].font = Font(
                    bold=True, name=DEFAULT_FONT.name
                )
            elif re.match(r"[RC][\d]+\.[\d]+|_HGNC:[\d]+", cell[0].value):
                sheet[cell[0].coordinate].font = Font(
                    bold=True, name=DEFAULT_FONT.name
                )
            if lock_cells:
                cell[0].protection = Protection(locked=True)

    # make hyperlinks blue
    # Lock any cells in column B for editing if they're populated'
    for cell in sheet.iter_rows(min_col=2, max_col=2):
        if cell[0].value:
            if "HYPERLINK" in str(cell[0].value):
                sheet[cell[0].coordinate].font = Font(
                    color="00007f", name=DEFAULT_FONT.name
                )
            if lock_cells:
                cell[0].protection = Protection(locked=True)

    output_file = f"{project_name}_{today}.xlsx"
    writer.book.save(output_file)

    print(f"Output written to {output_file}")

    return output_file
