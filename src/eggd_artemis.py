#!/usr/bin/env python
# eggd_artemis
import concurrent
import os
import pip
import re
import dxpy
import json
import datetime
import logging
from collections import defaultdict
from copy import deepcopy

# Install required packages
for package in os.listdir("/home/dnanexus/packages"):
    print(f"Installing {package}")
    pip.main(["install", "--no-index", "--no-deps", f"packages/{package}"])

from mergedeep import merge
from openpyxl.styles import DEFAULT_FONT, Font, Protection
import pandas as pd


def find_snv_files(reports):
    """Gather files related to SNV reports

    Args:
        reports (list): List of SNV report dxpy describe dicts

    Returns:
        snv_data (dict): Nested dictionary of files with sample name
            as key and list of SNV files within each clinical indication
    Example:
    {
        'X123456-GM1234567-23NGCEN23-1234-F-99347387: {
            'sample': 'X123456-GM1234567-23NGCEN23-1234-F-99347387',
            'Alignment BAI': {'$dnanexus_link: 'file-XX'},
            'Alignment BAI': {'$dnanexus_link: 'file-XY'},
            'clinical_indications': {
                'R414.1_APC associated Polyposis_G': {
                    'SNV': [
                        {
                            'SNV variant report': 'file-ZX',
                            'Coverage report': {'$dnanexus_link': 'file-GH'},
                            'SNV count': 1
                        }
                    ],
                'R208.1_Inherited breast cancer and ovarian cancer_P': {
                    'SNV': [
                        {
                            'SNV variant report': 'file-GH',
                            'Coverage report': {'$dnanexus_link': 'file-GH'},
                            'SNV count': 3
                        }
                    ]
                }
            }
        }
    }
    """
    def _find(report):
        """
        Find files for single sample

        Args:
            report (dict): dx describe return for single sample

        Returns:
            data (dict): files found for sample
        """
        # Get sample name
        sample = report['describe']['name'].split("_")[0]
        # Get file 'details' of the SNV xlsx report so that we can query
        # the clinical indication and variant count. If file has no 'details'
        # set default return to "Unknown"
        file_details = dxpy.DXFile(report['id']).get_details()
        clinical_indication = file_details.get(
            'clinical_indication', 'Unknown'
        )
        snv_variant_count = file_details.get('included', 'Unknown')

        # Get the job id that created the report
        job_id = report['describe']['createdBy']['job']

        # Get the workflow id that included the job
        parent_analysis = dxpy.bindings.dxjob.DXJob(
            dxid=job_id).describe()["parentAnalysis"]

        # Get the vcf file id and athena coverage file id
        parent_details = dxpy.bindings.dxanalysis.DXAnalysis(
            dxid=parent_analysis).describe()
        try:
            vcf_file = parent_details["input"]["stage-rpt_vep.vcf"]
        except KeyError:
            vcf_file = parent_details["input"]["stage-G9Q0jzQ4vyJ3x37X4KBKXZ5v.vcf"]
        try:
            coverage_report = parent_details["output"]["stage-rpt_athena.report"]
        except KeyError:
            coverage_report = parent_details["output"]["stage-Fyq5z18433GfYZbp3vX1KqjB.report"]

        # Extract the sention job id from the vcf metadata
        sention_job_id=dxpy.describe(vcf_file)["createdBy"]["job"]
        sentieon_details = dxpy.bindings.dxjob.DXJob(dxid=sention_job_id).describe()

        # Get bam & bai job id from sention job metadata
        mappings_bam = sentieon_details["output"]["mappings_bam"]
        mappings_bai = sentieon_details["output"]["mappings_bam_bai"]

        # Store in dictionary to return
        data = {
            "sample": sample,
            'Alignment BAM': mappings_bam,
            'Alignment BAI': mappings_bai,
            'clinical_indications': {
                clinical_indication: {
                    'SNV': [
                        {
                            "SNV variant report": report['describe']['id'],
                            "Coverage report": coverage_report,
                            "SNV count": str(snv_variant_count)
                        }
                    ]
                }
            }
        }

        return data

    snv_data = defaultdict(
        lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    )

    with concurrent.futures.ThreadPoolExecutor(max_workers=32) as executor:
        # submit jobs mapping each id to describe call
        concurrent_jobs = {
            executor.submit(_find, report) for report in reports
        }

        for future in concurrent.futures.as_completed(concurrent_jobs):
            # access returned output as each is returned in any order
            try:
                data = future.result()
                # Get info from the returned dict (only ever 1 clinical
                # indication and one set of SNV files)
                sample = data['sample']
                clin_ind = list(data['clinical_indications'])[0]
                snv_files = data['clinical_indications'][clin_ind]['SNV'][0]

                # Add in sample name, BAM and BAI info then append the SNV
                # file info to the relevant clinical indication
                snv_data[sample]['sample'] = data['sample']
                snv_data[sample]['Alignment BAM'] = data['Alignment BAM']
                snv_data[sample]['Alignment BAI'] = data['Alignment BAI']
                snv_data[sample][
                    'clinical_indications'
                ][clin_ind]['SNV'].append(snv_files)

            except Exception as exc:
                # catch any errors that might get raised during querying
                print(f"Error getting data for {concurrent_jobs[future]}: {exc}")

    return snv_data


def get_cnv_call_details(reports):
    """ Gets the job id of the cnv call job and stores input and output
        files in a dictionary for later use

    Args:
        reports (list): list of cnv report objects from dxpy search

    Returns:
        calling_files (str): dictionary storing all input and output information
    """
    # Get the job id of the generate workbook job
    gen_xlsx_job = reports[0]["describe"]["createdBy"]["job"]

    # Find the reports workflow analysis id
    reports_analysis = dxpy.bindings.dxjob.DXJob(
        dxid=gen_xlsx_job).describe()["parentAnalysis"]

    # Find the input vcf id
    try:
        vcf_id = dxpy.bindings.dxanalysis.DXAnalysis(
            dxid=reports_analysis).describe()["input"]["stage-cnv_vep.vcf"]
    except KeyError:
        vcf_id = dxpy.bindings.dxanalysis.DXAnalysis(
            dxid=reports_analysis).describe()["input"]["stage-GFYvJF04qq8VKgq34j30pZZ3.vcf"]

    # Find the cnv call job id
    cnv_call_job = dxpy.describe(vcf_id)["createdBy"]["job"]

    # Store all file ids and names in a dictionary
    calling_files = {}

    # Find the output files of the cnv call job
    cnv_details = dxpy.bindings.dxjob.DXJob(dxid=cnv_call_job).describe()
    gcnv_output = cnv_details["output"]["result_files"]
    gcnv_input = cnv_details["input"]["bambais"]

    # get and store name and file ID of all input and output files
    with concurrent.futures.ThreadPoolExecutor(max_workers=32) as executor:
        # submit jobs mapping each id to describe call
        concurrent_jobs = {
            executor.submit(
                dxpy.describe, file, fields={"name": True, "id": True}
            ) for file in gcnv_input + gcnv_output
        }

        for future in concurrent.futures.as_completed(concurrent_jobs):
            # access returned output as each is returned in any order
            try:
                data = future.result()
                calling_files[data['name']] = data['id']
            except Exception as exc:
                # catch any errors that might get raised during querying
                print(f"Error getting data for {concurrent_jobs[future]}: {exc}")

    print(f"Found {len(calling_files.keys())} gCNV input / output files")

    return calling_files


def get_cnv_file_ids(reports, gcnv_dict):
    """ Gather all related file ids for cnv files

    Args:
        reports (list): list of cnv report objects from dxpy search
        gcnv_dict(dict): dictionary output of gcnv call i/o files

    Returns:
        cnv_data (dict): Nested dictionary of files with sample name
            as key and nested dict of CNV files by clinical indication
    Example:
    {
        'X123456-GM1234567-23NGCEN23-1234-F-99347387: {
            'sample': 'X123456-GM1234567-23NGCEN23-1234-F-99347387',
            'Alignment BAI': {'$dnanexus_link: 'file-XX'},
            'Alignment BAI': {'$dnanexus_link: 'file-XY'},
            'clinical_indications': {
                'R414.1_APC associated Polyposis_G': {
                    'CNV': [
                        {
                            'CNV variant report': 'file-ZX',
                            'CNV calls for IGV': {'$dnanexus_link': 'file-TU'},
                            'CNV count': 0,
                            'Session file name': 'R414.1_CNV_1'
                        }
                    ],
                'R208.1_Inherited breast cancer and ovarian cancer_P': {
                    'CNV': [
                        {
                            'CNV variant report': 'file-OP',
                            'CNV calls for IGV': {'$dnanexus_link': 'file-LM'},
                            'CNV count': 0,
                            'Session file name': 'R208.1_CNV_1'
                        }
                    ]
                }
            },
            'CNV visualisation': 'file-ABC'
        }
    }
    """
    def _find(report):
        """
        Find files for single sample

        Args:
            report (dict): dx describe return for single sample

        Returns:
            data (dict): files found for sample
        """
        report_name = report['describe']['name']
        sample = report_name.split("_")[0]
        # Get end of CNV report name for naming IGV session file later
        # because if multiple panels exist for 1 sample we don't want to name
        # the separate session files the same
        cnv_file_ending = report_name.split('_', 1)[1].replace('.xlsx', '')

        # Get file 'details' of the CNV xlsx report so that we can query
        # the clinical indication and variant count. If file has no 'details'
        # set default return to "Unknown"
        file_details = dxpy.DXFile(report['id']).get_details()
        clin_ind = file_details.get(
            'clinical_indication', 'Unknown'
        )
        cnv_variant_count = file_details.get('variants', 'Unknown')

        gen_xlsx_job = report["describe"]["createdBy"]["job"]

        # Find the reports workflow analysis id
        reports_analysis = dxpy.bindings.dxjob.DXJob(
            dxid=gen_xlsx_job).describe()["parentAnalysis"]
        reports_details = dxpy.bindings.dxanalysis.DXAnalysis(
            dxid=reports_analysis).describe()

        # Get the CNV report and seg file using the analysis id
        try:
            cnv_workbook_id = reports_details["output"]["stage-cnv_generate_workbook.xlsx_report"]
        except KeyError:
            cnv_workbook_id = reports_details["output"]["stage-GFfYY9j4qq8ZxpFpP8zKG7G0.xlsx_report"]
        try:
            seg_id = reports_details["output"]["stage-cnv_additional_tasks.seg_file"]
        except KeyError:
            seg_id = reports_details["output"]["stage-GG2z5yQ4qq8vb2xp4pB8XByz.seg_file"]

        # gCNV job has all input bams and all outputs go through info
        # saved in the dictionary and find the sample specific files required
        for k, v in gcnv_dict.items():
            if k.startswith(f'{sample}'):
                if k.endswith(".bam"):
                    bam = v
                elif k.endswith(".bai"):
                    bai = v
                elif k.endswith(".bed.gz"):
                    gcnv_bed = v

        # Store in dictionary to return
        data = {
            'sample': sample,
            'Alignment BAM': bam,
            'Alignment BAI': bai,
            'CNV visualisation': gcnv_bed,
            'clinical_indications': {
                clin_ind: {
                    'CNV': [
                        {
                            'CNV variant report': cnv_workbook_id,
                            'CNV count': str(cnv_variant_count),
                            'CNV calls for IGV': seg_id,
                            'Session file name': cnv_file_ending
                        }
                    ]
                }
            }
        }

        return data


    cnv_data = defaultdict(
        lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    )

    with concurrent.futures.ThreadPoolExecutor(max_workers=32) as executor:
        # submit jobs mapping each id to describe call
        concurrent_jobs = {
            executor.submit(_find, report) for report in reports
        }

        for future in concurrent.futures.as_completed(concurrent_jobs):
            # access returned output as each is returned in any order
            try:
                data = future.result()

                # Get info from the returned dict (will only ever be one
                # clinical indication / CNV file dict)
                sample = data['sample']
                clin_ind = list(data['clinical_indications'])[0]
                cnv_info = data['clinical_indications'][clin_ind]['CNV'][0]

                # If sample and clinical indication already exists, append CNV
                # file info to existing list.
                # If sample doesn't exist yet just add the data as is
                cnv_data[sample]['sample'] = data['sample']
                cnv_data[sample]['Alignment BAM'] = data['Alignment BAM']
                cnv_data[sample]['Alignment BAI'] = data['Alignment BAI']
                cnv_data[sample][
                    'CNV visualisation'
                ] = data['CNV visualisation']
                cnv_data[sample][
                    'clinical_indications'
                ][clin_ind]['CNV'].append(cnv_info)

            except Exception as exc:
                # catch any errors that might get raised during querying
                print(f"Error getting data for {concurrent_jobs[future]}: {exc}")

    return cnv_data


def get_excluded_intervals(gcnv_output_dict):
    """ Get the excluded regions file from the gcnv dictionary

    Args:
        gcnv_output_dict (dict): dictionary of gcnv i/o files

    Returns:
        excluded_file (str): file id of the excluded regions file
    """
    for k, v in gcnv_output_dict.items():
        if k.endswith("excluded_intervals.bed"):
            return v
    return None


def get_multiqc_report(path_to_reports, project):
    """ Find appropriate multiqc file

    Args:
        snv_path (string): path to snv reports

    Returns:
        multiqc_file (str): file id of the multiqc file
    """
    # Get path to single from reports path
    single=f"/output/{path_to_reports.split('/')[2]}"

    # Find MultiQC jobs in the project
    multiqc_reports=list(dxpy.bindings.search.find_jobs(
            name_mode='glob',
            name="*MultiQC*",
            state="done",
            project=project,
            describe=True))

    # In case there's more than one MultiQC job
    # Find the one that has the exists in the single path in question
    for report in multiqc_reports:
        if report['describe']["folder"].startswith(single):
            multiqc = report["describe"]["output"]['multiqc_html_report']

    return multiqc


def make_url(file_id, project, url_duration):
    """ Given a file id create a download url

    Args:
        file_id (string/dict): dxpy file id or dnanexus file link
        project (string): id of project
        url_duration (int, optional): URL duration in seconds.

    Returns:
        file_url (string): Download url of requested file
    """
    # Bind dxpy file object to pass to make_download_url command
    file_info = dxpy.bindings.dxfile.DXFile(dxid=file_id, project=project)

    # Extract the file name to allow it to be used in the url
    file_name = file_info.describe()["name"]

    # Duration currently defaults to 28 days unless provided as input
    file_url = file_info.get_download_url(
            duration=url_duration, preauthenticated=True,
            filename=file_name, project=project)[0]

    # Without unsetting and clearing the workstation environment
    # the output urls have the local hostname which doesn't work
    # Find hostname
    workstation_hostname = file_url.split('/')[2]
    dx_hostname = 'dl.ec1.dnanex.us'

    # Replace with eu specific hostname and update protocol
    file_url = file_url.replace(workstation_hostname, dx_hostname)
    file_url = file_url.replace('http', 'https')

    return file_url


def set_order_map(snv_only=False):
    """ Set the order of the session depending on input

    Args:
        snv_only (bool, optional): If True, returns snv order map

    Returns:
        order_map (dict): order of inputs to the session template
    """

    if snv_only == True:

        order_map = {
            '6': {
                'url': '',
                'indexURL': ''
            },
            '7': {
                'url': '',
                'name': 'vcf'
            }
        }

    else:

        order_map = {
            '6': {
                'url': '',
                'name': 'CNV-bed'
            },
            '7': {
                'url': '',
                'indexURL': ''
            },
            '8': {
                'url': '',
                'name': 'variant-bed'
            },
            '9': {
                'url': '',
                'name': 'excluded-regions'
            },
            '10': {
                'url': '',
                'name': 'CNV-targets-bed'
            }

        }

    return order_map


def make_cnv_session(
    sample, session_file_name, bam_url, bai_url, bed_url,
    seg_url, excluded_url, targets_url, job_output, expiry_date
):
    """ Create a session file for IGV

    Args:
        sessions_list (list): list of session files to upload in the end
        sample (string): sample name
        session_file_name (string): str to add to end of session file name
        bam_url (string): URL of the bam file
        bai_url (string): URL of the bai file
        bed_url (string): URL of the bed file
        seg_url (string): URL of the seg file
        excluded_url (string): URL of the excluded regions file
        targets_url (string): URL of the targets file
        job_output (str) : output folder set for job
        expiry_date (string): date of expiry of file created

    Returns:
        session_file (string): IGV session file URL
    """
    order_map = set_order_map()

    # Copy template to avoid overwriting
    with open('/home/dnanexus/cnv-template.json') as fh:
        template = json.load(fh)

    template_copy = deepcopy(template)

    template_copy['tracks'] = []

    # Order 6 - Visualisation bed for CNV calls
    order_map['6']['url'] = bed_url

    # Order 7 - Patient bam and index
    order_map['7']['name'] = sample
    order_map['7']['url'] = bam_url
    order_map['7']['indexURL'] = bai_url

    # Order 8 - Variant seg file
    order_map['8']['url'] = seg_url

    # Order 9 - Excluded regions from CNV calling bed
    order_map['9']['url'] = excluded_url

    # Order 10 - Target regions bed
    order_map['10']['url'] = targets_url

    # Map urls to template
    for track in template['tracks']:
        if track['order'] in [6, 7, 8, 9, 10]:
            order = str(track['order'])
            track['url'] = order_map[order]['url']

        if track['order'] == 7:
            track['indexURL'] = order_map[order]['indexURL']

        template_copy['tracks'].append(track)

    output_name = f"{sample}_{session_file_name}_igv.json"

    with open(output_name, "w") as outfile:
        json.dump(template_copy, outfile, indent=4)

    session_file = dxpy.upload_local_file(
        output_name,
        folder=f'{job_output}/igv_sessions',
        tags=[expiry_date],
        wait_on_close=True)

    session_file_id = session_file.get_id()

    return session_file_id


def generate_sample_urls(
    sample, sample_data, url_duration, ex_intervals_url,
    bed_url, expiry_date, job_output) -> dict:
    """
    Generates all URLs and session file for a given sample

    Parameters
    ----------
    sample : str
        sample name
    sample_data : dict
        file IDs of sample related files
    url_duration : int
        URL duration in seconds
    ex_intervals_url : str
        URL of excluded intervals file
    bed_url : str
        URL of bed file
    expiry_date : str
        date of URL expiration
    job_output : str
        output folder set for job

    Returns
    -------
    urls : dict
        dict of sample URLs
    """
    dx_project = os.environ.get("DX_PROJECT_CONTEXT_ID")

    urls = defaultdict(lambda: defaultdict(dict))
    urls['sample'] = sample

    bam_url = make_url(sample_data['Alignment BAM'], dx_project, url_duration)
    bai_url = make_url(sample_data['Alignment BAI'], dx_project, url_duration)

    urls['bam_url'] = bam_url
    urls['bai_url'] = bai_url

    # Loop over clinical indications
    for clin_ind in sample_data['clinical_indications']:
        snv_files = sample_data['clinical_indications'][clin_ind].get('SNV')
        # If we have SNV reports, for each report make URLs and append to
        # SNV value
        if snv_files:
            urls['clinical_indications'][clin_ind]['SNV'] = []

            for snv_file in snv_files:
                coverage_url = make_url(
                    snv_file['Coverage report'], dx_project, url_duration
                )
                snv_url = make_url(
                    snv_file['SNV variant report'], dx_project, url_duration
                )

                urls['clinical_indications'][clin_ind]['SNV'].append(
                    {
                        'SNV count': snv_file['SNV count'],
                        'coverage_url': (
                            f'=HYPERLINK("{coverage_url}", "{coverage_url}")'
                        ),
                        'snv_url': f'=HYPERLINK("{snv_url}", "{snv_url}")'
                    }
                )

        # If we have CNV reports, for each report make URLs/session files
        # and append to CNV value
        cnv_files = sample_data['clinical_indications'][clin_ind].get('CNV')
        if cnv_files:
            cnv_bed = make_url(
                sample_data['CNV visualisation'], dx_project, url_duration
            )
            urls['clinical_indications'][clin_ind]['CNV'] = []

            for cnv_file in cnv_files:
                session_file_end = cnv_file['Session file name']
                cnv_url = make_url(
                    cnv_file['CNV variant report'], dx_project, url_duration
                )
                cnv_seg = make_url(
                    cnv_file['CNV calls for IGV'], dx_project, url_duration
                )

                cnv_session = make_cnv_session(
                    sample, session_file_end, bam_url, bai_url, cnv_bed,
                    cnv_seg, ex_intervals_url, bed_url, job_output,
                    expiry_date
                )

                cnv_session_url = make_url(
                    cnv_session, dx_project, url_duration
                )

                urls['clinical_indications'][clin_ind]['CNV'].append({
                    'CNV count': cnv_file['CNV count'],
                    'cnv_bed': cnv_bed,
                    'cnv_seg': cnv_seg,
                    'cnv_session_fileid': cnv_session,
                    'cnv_url': f'=HYPERLINK("{cnv_url}", "{cnv_url}")',
                    'cnv_session_url': (
                        f'=HYPERLINK("{cnv_session_url}", "{cnv_session_url}")'
                    )
                })

    return urls


def remove_url_if_variant_count_is_zero(all_sample_urls):
    """
    Remove links to download Excel reports if there are no variants

    Parameters
    ----------
    all_sample_urls : dict
        generated URLs and file metadata for each sample

    Returns
    -------
    all_sample_urls: dict
        generated URLs and file metadata for each sample with Excel
        download URLs removed when no variants are present
    """
    for sample, sample_data in all_sample_urls.items():
        for clin_ind in sample_data['clinical_indications']:
            file_data = sample_data['clinical_indications'][clin_ind]

            # Replace SNV report URL with text if variant count is zero.
            # If variant count was never present then replace with
            # None so we can easily not include this field in final .xlsx
            if 'SNV' in file_data:
                for file in file_data['SNV']:
                    if file.get('SNV count') == '0':
                        file['snv_url'] = 'No SNVs post-filtering'
                    elif file.get('SNV count') == 'Unknown':
                        file['SNV count'] = None

            # Do same for CNVs
            if 'CNV' in file_data:
                for file in file_data['CNV']:
                    if file.get('CNV count') == '0':
                        file['cnv_url'] = 'No CNVs detected'
                    elif file.get('CNV count') == 'Unknown':
                        file['CNV count'] = None

    return all_sample_urls


def write_output_file(
    sample_urls, today, expiry_date, multiqc_url, qc_url, project_name,
    lock_cells
):
    """
    Writes output xlsx file with all download URLs

    Parameters
    ----------
    sample_urls : dict
        generated URLs for each sample
    today : str
        today date
    expiry_date : str
        date of URL experation
    multiqc_url : str
        download URL for multiQC report
    qc_url : str
        download URL for QC report
    project_name: str
        DNAnexus project name
    lock_cells : boolean
        determines whether to lock any populated cells for editing

    Outputs
    -------
    xlsx file
    """
    print("Writing output file")

    sample_count = str(len(sample_urls.keys()))

    multiqc_url = f'=HYPERLINK("{multiqc_url}", "{multiqc_url}")'
    if qc_url.startswith('http'):
        qc_url = f'=HYPERLINK("{qc_url}", "{qc_url}")'

    df = pd.DataFrame(columns=['a', 'b'])
    df = df.append({'a': 'Run:', 'b': project_name}, ignore_index=True)
    df = df.append(
        {'a': 'Number of samples in this file:', 'b': sample_count},
        ignore_index=True
    )
    df = df.append({}, ignore_index=True)
    df = df.append({'a': 'Date Created:', 'b': today}, ignore_index=True)
    df = df.append({'a': 'Expiry Date:', 'b': expiry_date}, ignore_index=True)
    df = df.append({}, ignore_index=True)
    df = df.append({'a': 'Run Level Files'}, ignore_index=True)
    df = df.append({'a': 'MultiQC report', 'b': multiqc_url}, ignore_index=True)
    df = df.append({'a': 'QC Status Report', 'b': qc_url}, ignore_index=True)
    df = df.append({}, ignore_index=True)
    df = df.append({}, ignore_index=True)
    df = df.append({'a': 'Per Sample Files'}, ignore_index=True)
    df = df.append({}, ignore_index=True)

    sample_order = sorted(sample_urls.keys())

    for sample in sample_order:
        urls = sample_urls.get(sample)

        df = df.append({'a': sample}, ignore_index=True)

        # Fields we need once per sample
        sample_level_urls = {
            'bam_url': 'Alignment BAM',
            'bai_url': 'Alignment BAI'
        }

        # Fields we need once per report
        url_fields = {
            'coverage_url': 'Coverage report',
            'SNV count': 'SNV count post-filtering',
            'snv_url': 'Small variant report',
            'CNV count': 'CNV count',
            'cnv_url': 'CNV variant report',
            'cnv_session_url': 'CNV IGV Session'
        }

        for clinical_indication in urls['clinical_indications']:
            file_data = urls['clinical_indications'][clinical_indication]

            # If we know clinical indication add this as a header
            if clinical_indication != 'Unknown':
                df = df.append({'a': clinical_indication}, ignore_index=True)

            # Add fields for any SNV files first for that clinical indication
            for field, label in url_fields.items():
                for snv_data in file_data.get('SNV', {}):
                    if snv_data.get(field):
                        df = df.append(
                            {'a': label, 'b': snv_data.get(field)},
                            ignore_index=True
                        )

                # Then add fields for CNV fields for that clinical indication
                for cnv_data in file_data.get('CNV', {}):
                    if cnv_data.get(field):
                        df = df.append(
                            {'a': label, 'b': cnv_data.get(field)},
                            ignore_index=True
                        )
        # Add BAM and BAI URLs for the sample
        for field, label in sample_level_urls.items():
            if urls.get(field):
                if field == 'bam_url':
                    df = df.append({}, ignore_index=True)

                df = df.append(
                    {'a': label, 'b': urls.get(field)}, ignore_index=True
                )

        df = df.append({}, ignore_index=True)
        df = df.append({}, ignore_index=True)


    writer = pd.ExcelWriter(f'{project_name}_{today}.xlsx', engine='openpyxl')
    df.to_excel(writer, index=False, header=False)

    # set column widths
    sheet = writer.sheets['Sheet1']
    sheet.column_dimensions['A'].width = 55
    sheet.column_dimensions['B'].width = 155

    sheet['A1'].font = Font(bold=True, name=DEFAULT_FONT.name)
    sheet['A2'].font = Font(bold=True, name=DEFAULT_FONT.name)
    sheet['A7'].font = Font(bold=True, name=DEFAULT_FONT.name)
    sheet['A12'].font = Font(bold=True, name=DEFAULT_FONT.name)

    # If lock_cells=True we're protecting populated cells from editing
    # openpyxl is silly and you need to lock a whole sheet then unlock
    # specific cells - so to do the opposite we unlock the first 3000 rows
    # and 100 columns (assuming max 20 rows / sample and 96 samples per run).
    # Also set format so that we can still resize rows/columns if protected
    if lock_cells:
        print("lock_cells=True, locking any populated Excel cells")
        sheet.protection.sheet = True
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False
        sheet.protection.formatCells = False

        for row_no in range(1, 3000):
            for col_no in range(1, 100):
                sheet.cell(row=row_no, column=col_no).protection = Protection(
                    locked=False
                )
    else:
        print("lock_cells=False, all Excel cells will be editable")

    # Make sample IDs and any clinical indication names bold
    # Lock any cells in column A for editing if they're populated
    for cell in sheet.iter_rows(max_col=1):
        if cell[0].value:
            if re.match(r'X[\d]+', cell[0].value):
                sheet[cell[0].coordinate].font = Font(
                    bold=True, name=DEFAULT_FONT.name)
            elif re.match(r'[a-zA-Z0-9]+-[a-zA-Z0-9]+-[a-zA-Z0-9]+-[a-zA-Z0-9]+-[MFU]-[a-zA-Z0-9]+', cell[0].value):
                sheet[cell[0].coordinate].font = Font(
                    bold=True, name=DEFAULT_FONT.name)
            elif re.match(r'[RC][\d]+\.[\d]+|_HGNC:[\d]+', cell[0].value):
                sheet[cell[0].coordinate].font = Font(
                    bold=True, name=DEFAULT_FONT.name)
            if lock_cells:
                cell[0].protection = Protection(locked=True)

    # make hyperlinks blue
    # Lock any cells in column B for editing if they're populated'
    for cell in sheet.iter_rows(min_col=2, max_col=2):
        if cell[0].value:
            if 'HYPERLINK' in str(cell[0].value):
                sheet[cell[0].coordinate].font = Font(
                    color='00007f', name=DEFAULT_FONT.name)
            if lock_cells:
                cell[0].protection = Protection(locked=True)

    writer.book.save(f'{project_name}_{today}.xlsx')

    print(f"Output written to {project_name}_{today}.xlsx")


@dxpy.entry_point('main')
def main(
    url_duration, lock_cells=True, snv_path=None, cnv_path=None,
    bed_file=None, qc_status=None, multiqc_report=None
):

    # Set up logging
    logger = logging.getLogger(__name__)
    logger.addHandler(dxpy.DXLogHandler())
    logger.propagate = False
    logger.setLevel(logging.DEBUG)

    # Get the project ID
    DX_PROJECT = os.environ.get("DX_PROJECT_CONTEXT_ID")
    print(DX_PROJECT)

    # Set the environment context to allow upload
    dxpy.set_workspace_id(DX_PROJECT)

    # Get output folder set for this job
    job_output = dxpy.bindings.dxjob.DXJob(
        os.environ.get('DX_JOB_ID')).describe()['folder']

    # Make output folder for job
    dxpy.api.project_new_folder(
        DX_PROJECT,
        input_params={
            "folder": job_output,
            "parents": True})

    # Get name of project for output naming
    project_name = dxpy.describe(DX_PROJECT)['name']
    project_name = '_'.join(project_name.split('_')[1:-1])

    # Gather required SNV files if SNV path is provided
    if snv_path:
        logger.info("Gathering Small variant files")

        snv_data = {}

        for path in snv_path.split(','):
            print(f'Gathering reports from: {path}')
            snv_reports = list(dxpy.bindings.search.find_data_objects(
                name="*xlsx",
                name_mode='glob',
                folder=path,
                project=DX_PROJECT,
                describe=True))

            # Get SNV ids
            snv_files = find_snv_files(snv_reports)
            merge(snv_data, snv_files)

        print(f"Size of snv data dict: {len(snv_data.keys())}")

        # If multiqc report not given, search for multiqc report
        if not multiqc_report:
            print(
                "No MultiQC report given, searching instead "
                f"with path {snv_path}"
            )
            multiqc_report = get_multiqc_report(
                snv_path.split(',')[0], DX_PROJECT
            )

    # Gather required CNV files if CNV path is provided
    if cnv_path:
        logger.info("Gathering CNV files")

        # Create folder for session files
        dxpy.api.project_new_folder(
            DX_PROJECT,
            input_params={
                "folder": f"{job_output}/igv_sessions",
                "parents": True})

        cnv_data = {}

        for path in cnv_path.split(','):
            print(f'Gathering reports from: {path}')
            cnv_reports = list(dxpy.bindings.search.find_data_objects(
                name="*xlsx",
                name_mode='glob',
                folder=path,
                project=DX_PROJECT,
                describe=True))

            gcnv_job_info = get_cnv_call_details(cnv_reports)
            cnv_files = get_cnv_file_ids(cnv_reports, gcnv_job_info)

            # Get excluded intervals file
            excluded_intervals = get_excluded_intervals(gcnv_job_info)
            ex_intervals_url = make_url(
                excluded_intervals, DX_PROJECT, url_duration)

            merge(cnv_data, cnv_files)

        print(f"Size of cnv data dict: {len(cnv_data.keys())}")

        # If multiqc report not given or found earlier, search for it
        if not multiqc_report:
            print(
                "No MultiQC report given, searching instead "
                f"with path {cnv_path}"
            )
            multiqc_report = get_multiqc_report(
                cnv_path.split(',')[0], DX_PROJECT
            )
    else:
        ex_intervals_url = ''

    logger.info("Making URLs for additional files")

    # If a bed file is provided, add to a link to the output
    if bed_file:
        bed_file_url = make_url(
            bed_file, 'project-Fkb6Gkj433GVVvj73J7x8KbV', url_duration)
    else:
        # Setting as empty to avoid session errors
        bed_file_url = ''

    # If a QC Status xlsx is provided, add to a link to the output
    if qc_status:
        qc_status_url = make_url(qc_status, DX_PROJECT, url_duration)
    else:
        qc_status_url = 'No QC status file provided'

    file_data = {}

    if snv_path and cnv_path:
        merge(file_data, snv_data, cnv_data)
    elif snv_path:
        file_data = snv_data
    elif cnv_path:
        file_data = cnv_data
    else:
        logger.debug("No paths given, exiting...")
        exit(1)


    today = datetime.datetime.now().strftime("%Y-%m-%d")
    expiry_date=(
        datetime.datetime.strptime(datetime.datetime.now().strftime('%Y-%m-%d'),
        '%Y-%m-%d') + datetime.timedelta(seconds=url_duration)
    ).strftime('%Y-%m-%d')


    # generate all urls for each sample
    logger.info("Generating per sample URLs")
    all_sample_urls = {}

    with concurrent.futures.ThreadPoolExecutor(max_workers=32) as executor:
        # submit jobs mapping each id to describe call
        concurrent_jobs = {
            executor.submit(
                generate_sample_urls, sample, sample_data, url_duration,
                ex_intervals_url, bed_file_url, expiry_date, job_output
            ) for sample, sample_data in file_data.items()
        }

        for future in concurrent.futures.as_completed(concurrent_jobs):
            # access returned output as each is returned in any order
            try:
                data = future.result()
                all_sample_urls[data['sample']] = data
            except Exception as exc:
                # catch any errors that might get raised during querying
                print(f"Error getting data for {concurrent_jobs[future]}: {exc}")

    multiqc_url = make_url(multiqc_report, DX_PROJECT, url_duration)

    # Remove download URLs for reports with no variants in
    all_sample_urls = remove_url_if_variant_count_is_zero(all_sample_urls)

    write_output_file(
        all_sample_urls, today, expiry_date, multiqc_url,
        qc_status_url, project_name, lock_cells
    )

    # Upload output to the platform
    output = {}
    url_file = dxpy.upload_local_file(
        f'{project_name}_{today}.xlsx',
        folder=job_output,
        tags=[expiry_date]
    )
    output["url_file"] = dxpy.dxlink(url_file)

    # Find session files to link to output
    """
    Example all_sample_urls format we're looping over:
    'sample-name': {
        'sample': 'sample-name',
        'bam_url': 'url_for_bam',
        'bai_url':  'url_for_bai',
        'clinical_indications': {
            'R141.1': {
                'SNV': [{
                    'SNV count': '7',
                    'coverage_url: '=HYPERLINK("hyperlink", "hyperlink"),
                    'snv_url': '=HYPERLINK("hyperlink", "hyperlink"),
                }],
                'CNV': [{
                    'CNV count': '1',
                    'cnv_bed': 'bed_url',
                    'cnv_seg': 'seg_url',
                    'cnv_session_fileid': 'file-XYZ',
                    'cnv_url': '=HYPERLINK("hyperlink", "hyperlink"),
                    'cnv_session_url': '=HYPERLINK("hyperlink", "hyperlink")
                }]
            }
        }
    }
    """
    session_files = []
    for sample, sample_info in all_sample_urls.items():
        for clin_ind in sample_info.get('clinical_indications'):
            if sample_info['clinical_indications'][clin_ind].get('CNV'):
                for cnv_item in (
                    sample_info['clinical_indications'][clin_ind]['CNV']
                ):
                    session_file_id = cnv_item['cnv_session_fileid']
                    session_files.append(session_file_id)

    print(f"Found {len(session_files)} session files to link to job output")
    if session_files:
        output["session_files"] = [dxpy.dxlink(item) for item in session_files]


    return output

dxpy.run()
